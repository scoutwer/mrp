"""
excel_export.py — Profesjonalny eksport danych MRP do pliku Excel.

Moduł generuje sformatowany raport Excel z użyciem openpyxl:
- Obramowania (cienkie szare linie)
- Kolorystyka nagłówków (ciemny granat, biały tekst)
- Warunkowe formatowanie wiersza "Przewidywane na stanie"
- Ukrywanie zer (puste komórki)
- Stopka z parametrami
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ==========================================================================
# Stałe stylów
# ==========================================================================

# Obramowanie — cienkie szare linie
_THIN_BORDER = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)

# Nagłówek tabeli (wiersz z numerami tygodni)
_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# Pierwsza kolumna nagłówka ("Dane produkcyjne" / "Okres")
_HEADER_FIRST_COL_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

# Kolumna etykiet (nazwy wierszy)
_LABEL_FONT = Font(name="Segoe UI", size=10)
_LABEL_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
_LABEL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

# Komórki danych (normalne)
_DATA_FONT = Font(name="Segoe UI", size=10)
_DATA_ALIGN = Alignment(horizontal="center", vertical="center")
_DATA_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Wiersz "Przewidywane na stanie" — normalny (>=0)
_AVAIL_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
_AVAIL_FONT = Font(name="Segoe UI", size=10, bold=True, color="1A5276")
_AVAIL_LABEL_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
_AVAIL_LABEL_FONT = Font(name="Segoe UI", size=10, bold=True, color="1A5276")

# Wiersz "Przewidywane na stanie" — wartość < 0
_AVAIL_NEGATIVE_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
_AVAIL_NEGATIVE_FONT = Font(name="Segoe UI", size=10, bold=True, color="C0392B")

# Stopka — etykiety i wartości
_FOOTER_LABEL_FONT = Font(name="Segoe UI", size=10, bold=True, color="2C3E50")
_FOOTER_VALUE_FONT = Font(name="Segoe UI", size=10, color="333333")
_FOOTER_FILL = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")

# Szerokości kolumn
_FIRST_COL_WIDTH = 33
_DATA_COL_WIDTH = 10


# ==========================================================================
# Nazwy wierszy tabeli MRP
# ==========================================================================

_NAZWY_WIERSZY = [
    "Całkowite zapotrzebowanie",
    "Planowane przyjęcia",
    "Przewidywane na stanie",
    "Zapotrzebowanie netto",
    "Planowane zamówienia",
    "Planowane przyjęcie zamówień",
]

_WIERSZ_AVAIL = "Przewidywane na stanie"


# ==========================================================================
# Główna funkcja eksportu
# ==========================================================================

def eksportuj_excel(dane_mrp: dict, sciezka_pliku: str) -> None:
    """
    Generuje profesjonalnie sformatowany raport MRP w formacie Excel (.xlsx).

    Args:
        dane_mrp: Słownik z danymi MRP (klucze: nazwa, lt, partia, typ_bom,
                  zapas, rodzic, mnoznik, gross, sched, avail, net, rel, rec, n).
        sciezka_pliku: Ścieżka do pliku wyjściowego .xlsx.

    Raises:
        PermissionError: Jeśli plik jest otwarty w innej aplikacji.
        IOError: Jeśli zapis się nie powiedzie.
    """
    n = dane_mrp["n"]
    nazwa = dane_mrp.get("nazwa", "MRP")

    # Przygotuj dane wierszy
    dane_wierszy = {
        "Całkowite zapotrzebowanie": dane_mrp["gross"],
        "Planowane przyjęcia": dane_mrp["sched"],
        "Przewidywane na stanie": dane_mrp["avail"],
        "Zapotrzebowanie netto": dane_mrp["net"],
        "Planowane zamówienia": dane_mrp["rel"],
        "Planowane przyjęcie zamówień": dane_mrp["rec"],
    }

    wb = Workbook()
    # Nazwa arkusza — max 31 znaków (limit Excela)
    sheet_name = nazwa[:31] if nazwa else "Arkusz1"
    ws = wb.active
    ws.title = sheet_name

    # ------------------------------------------------------------------
    # 1. Nagłówek tabeli (wiersz 1)
    # ------------------------------------------------------------------
    row_idx = 1

    # Komórka A1 — "Dane produkcyjne" / "Okres"
    cell = ws.cell(row=row_idx, column=1, value="Dane produkcyjne")
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FIRST_COL_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _THIN_BORDER

    # Komórki B1..N1 — numery tygodni
    for i in range(n):
        col = i + 2
        cell = ws.cell(row=row_idx, column=col, value=i + 1)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # ------------------------------------------------------------------
    # 2. Wiersze danych (wiersze 2..7)
    # ------------------------------------------------------------------
    for r, nazwa_wiersza in enumerate(_NAZWY_WIERSZY):
        row_idx = r + 2
        wartosci = dane_wierszy[nazwa_wiersza]
        jest_avail = (nazwa_wiersza == _WIERSZ_AVAIL)

        # Kolumna A — etykieta wiersza
        cell = ws.cell(row=row_idx, column=1, value=nazwa_wiersza)
        cell.border = _THIN_BORDER

        if jest_avail:
            cell.font = _AVAIL_LABEL_FONT
            cell.fill = _AVAIL_LABEL_FILL
        else:
            cell.font = _LABEL_FONT
            cell.fill = _LABEL_FILL
        cell.alignment = _LABEL_ALIGN

        # Kolumny danych
        for i in range(n):
            col = i + 2
            wartosc = wartosci[i] if i < len(wartosci) else 0

            # Formatowanie wartości wyświetlanej
            if jest_avail:
                # "Przewidywane na stanie" — zawsze pokazuj wartość (nawet 0)
                display_val = wartosc
            else:
                # Pozostałe wiersze — ukryj zera
                display_val = wartosc if wartosc != 0 else ""

            cell = ws.cell(row=row_idx, column=col, value=display_val)
            cell.border = _THIN_BORDER
            cell.alignment = _DATA_ALIGN

            if jest_avail:
                if wartosc < 0:
                    cell.font = _AVAIL_NEGATIVE_FONT
                    cell.fill = _AVAIL_NEGATIVE_FILL
                else:
                    cell.font = _AVAIL_FONT
                    cell.fill = _AVAIL_FILL
            else:
                cell.font = _DATA_FONT
                cell.fill = _DATA_FILL

    # ------------------------------------------------------------------
    # 3. Stopka z parametrami (2 wiersze pod tabelą)
    # ------------------------------------------------------------------
    stopka_row1 = len(_NAZWY_WIERSZY) + 3  # +1 nagłówek, +1 przerwa
    stopka_row2 = stopka_row1 + 1

    # Parametry w stopce
    _zapisz_parametr_stopki(ws, stopka_row1, 1, "Czas realizacji", str(dane_mrp["lt"]))
    _zapisz_parametr_stopki(ws, stopka_row1, 3, "Wielkość partii", str(dane_mrp["partia"]))
    _zapisz_parametr_stopki(ws, stopka_row1, 5, "Na stanie", str(dane_mrp["zapas"]))

    # Poziom BOM — wyciągnij numer z tekstu typu "Wyrób gotowy (Poziom 0)"
    typ_bom = dane_mrp.get("typ_bom", "")
    poziom_bom = _wyciagnij_poziom_bom(typ_bom)
    _zapisz_parametr_stopki(ws, stopka_row2, 1, "Poziom BOM", str(poziom_bom))

    rodzic = dane_mrp.get("rodzic", "")
    if rodzic:
        _zapisz_parametr_stopki(ws, stopka_row2, 3, "Rodzic BOM", rodzic)
        mnoznik = dane_mrp.get("mnoznik", 1)
        _zapisz_parametr_stopki(ws, stopka_row2, 5, "Ilość w BOM", str(mnoznik))

    # ------------------------------------------------------------------
    # 4. Ustawienie szerokości kolumn
    # ------------------------------------------------------------------
    ws.column_dimensions["A"].width = _FIRST_COL_WIDTH

    for i in range(n):
        col_letter = get_column_letter(i + 2)
        ws.column_dimensions[col_letter].width = _DATA_COL_WIDTH

    # ------------------------------------------------------------------
    # 5. Zapis pliku
    # ------------------------------------------------------------------
    wb.save(sciezka_pliku)


# ==========================================================================
# Funkcje pomocnicze
# ==========================================================================

def _zapisz_parametr_stopki(ws, row: int, col: int, etykieta: str, wartosc: str) -> None:
    """Zapisuje parę etykieta + wartość w stopce z odpowiednim formatowaniem."""
    cell_label = ws.cell(row=row, column=col, value=f"{etykieta} =")
    cell_label.font = _FOOTER_LABEL_FONT
    cell_label.fill = _FOOTER_FILL
    cell_label.alignment = Alignment(horizontal="right", vertical="center")

    cell_value = ws.cell(row=row, column=col + 1, value=wartosc)
    cell_value.font = _FOOTER_VALUE_FONT
    cell_value.fill = _FOOTER_FILL
    cell_value.alignment = Alignment(horizontal="left", vertical="center")


def _wyciagnij_poziom_bom(typ_bom: str) -> int:
    """Wyciąga numer poziomu BOM z tekstu comboboxa, np. 'Wyrób gotowy (Poziom 0)' → 0."""
    import re
    match = re.search(r"Poziom\s+(\d+)", typ_bom)
    if match:
        return int(match.group(1))
    return 0
